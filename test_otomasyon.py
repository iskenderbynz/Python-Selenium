from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
from datetime import datetime
import openpyxl
from constants import globalConstants

class Test_Otomasyon:

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()


    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data

    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys(username)
        inputPass.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}-{self.whatTime()}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def test_invalid_login_username(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("")
        inputPass.send_keys("")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login_username-{self.whatTime()}.png")
        assert errorMessage.text == "Epic sadface: Username is required"

    def test_invalid_login_pass(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("standard_user")
        inputPass.send_keys("")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login_pass-{self.whatTime()}.png")
        assert errorMessage.text == "Epic sadface: Password is required"
        
    def test_invalid_login_locked(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("locked_out_user")
        inputPass.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login_locked-{self.whatTime()}.png")
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
    
    def test_invalid_login_namepass(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("")
        inputPass.send_keys("")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        inputIcon = self.driver.find_elements(By.XPATH,"//*[name()='svg']")
        if len(inputIcon) == 3:
            errorMessageClose = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3/button")
            errorMessageClose.click()
            inputIcon = self.driver.find_elements(By.XPATH,"//*[name()='svg']")
        testResult = len(inputIcon) == 0
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login_namepass-{self.whatTime()}.png")
        assert testResult
    
    def test_valid_login(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("standard_user")
        inputPass.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        testResult = self.driver.current_url == f"{globalConstants.URL}inventory.html"
        self.driver.save_screenshot(f"{self.folderPath}/test_valid_login-{self.whatTime()}.png")
        assert testResult
    
    def test_product_count(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("standard_user")
        inputPass.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        productCount = self.driver.find_elements(By.CLASS_NAME,"inventory_item_name")
        testResult = len(productCount) == 6
        self.driver.save_screenshot(f"{self.folderPath}/test_product_count-{self.whatTime()}.png")
        assert testResult

    def test_product_add_cart(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("standard_user")
        inputPass.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.waifForElementVisible((By.ID,"add-to-cart-sauce-labs-backpack"))
        product1Add = self.driver.find_element(By.ID,"add-to-cart-sauce-labs-backpack")
        product2Add = self.driver.find_element(By.ID,"add-to-cart-sauce-labs-bike-light")
        product1Add.click()
        product2Add.click()
        cartBtn = self.driver.find_element(By.XPATH,"/html/body/div/div/div/div[1]/div[1]/div[3]/a")
        cartBtn.click()
        cartProduct = self.driver.find_elements(By.CLASS_NAME,"inventory_item_name")
        testResult = len(cartProduct) == 2
        self.driver.save_screenshot(f"{self.folderPath}/test_product_add_cart-{self.whatTime()}.png")
        assert testResult

    def test_product_del_cart(self):
        self.waifForElementVisible((By.ID,"user-name"))
        inputUser = self.driver.find_element(By.ID,"user-name")
        self.waifForElementVisible((By.ID,"password"),10)
        inputPass = self.driver.find_element(By.ID,"password")
        inputUser.send_keys("standard_user")
        inputPass.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.waifForElementVisible((By.ID,"add-to-cart-sauce-labs-backpack"))
        product1Add = self.driver.find_element(By.ID,"add-to-cart-sauce-labs-backpack")
        product2Add = self.driver.find_element(By.ID,"add-to-cart-sauce-labs-bike-light")
        product1Add.click()
        product2Add.click()
        cartBtn = self.driver.find_element(By.XPATH,"/html/body/div/div/div/div[1]/div[1]/div[3]/a")
        cartBtn.click()
        product1Del = self.driver.find_element(By.ID,"remove-sauce-labs-backpack")
        product2Del = self.driver.find_element(By.ID,"remove-sauce-labs-bike-light")
        product1Del.click()
        product2Del.click()
        cartProduct = self.driver.find_elements(By.CLASS_NAME,"inventory_item_name")
        testResult = len(cartProduct) == 0
        self.driver.save_screenshot(f"{self.folderPath}/test_product_del_cart-{self.whatTime()}.png")
        assert testResult


    def waifForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(expected_conditions.visibility_of_element_located(locator))

    def whatTime(self):
        now = datetime.now()
        return  now.strftime("%H%M%S")
